from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from .models import Pilares, EquipoNoFuncional
import json
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER
import io

# ============ FUNCIÓN PARA VERIFICAR SI ES ADMIN ============
def es_admin(user):
    return user.is_authenticated and user.is_staff

# ============ VISTAS DE LOGIN/LOGOUT ============
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', '/')
            return redirect(next_url)
        else:
            return render(request, 'pilares/login.html', {'error': 'Usuario o contraseña incorrectos'})
    return render(request, 'pilares/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    """Vista principal del dashboard con estadísticas y tabla"""
    pilares = Pilares.objects.all()
    
    # Datos para la tabla
    pilares_data = []
    for p in pilares:
        pilares_data.append({
            'clave_id': p.clave_id,
            'nombre': p.nombre,
            'alcaldia': p.alcaldia,
            'zona': p.zona,
            'modelo_equipo': p.modelo_equipo,  # ← NUEVO
            'pasta_termica': p.pasta_termica,
            'mouse_nuevos': p.mouse_nuevos,
            'teclados_nuevos': p.teclados_nuevos,
            'equipos_8gb': p.equipos_8gb,
            'total_equipos': p.total_equipos,
            'equipos_inactivos': p.equipos_inactivos,  # ← NUEVO
            'porcentaje_8gb': p.porcentaje_8gb,
        })
    
    # Datos para gráficas por zona
    zonas = ['NORTE', 'SUR', 'PONIENTE', 'CENTRO', 'ORIENTE', 'SUR PONIENTE']
    zonas_data = []
    for zona in zonas:
        pilares_zona = [p for p in pilares if p.zona == zona]
        count = len(pilares_zona)
        equipos_8gb = sum(p.equipos_8gb for p in pilares_zona)
        total_equipos = sum(p.total_equipos for p in pilares_zona)
        porcentaje = round((equipos_8gb / total_equipos * 100), 1) if total_equipos > 0 else 0
        zonas_data.append({
            'nombre': zona,
            'count': count,
            'porcentaje': porcentaje,
        })
    
    # Estadísticas generales
    total_pilares = pilares.count()
    total_equipos_8gb = sum(p.equipos_8gb for p in pilares)
    total_equipos = sum(p.total_equipos for p in pilares)
    porcentaje_general = round((total_equipos_8gb / total_equipos * 100), 1) if total_equipos > 0 else 0
    
    context = {
        'pilares_data_json': json.dumps(pilares_data),
        'zonas_data_json': json.dumps(zonas_data),
        'total_pilares': total_pilares,
        'total_equipos_8gb': total_equipos_8gb,
        'total_equipos': total_equipos,
        'porcentaje_general': porcentaje_general,
        'es_admin': request.user.is_superuser,  # ejemplo
    }
    
    return render(request, 'pilares/dashboard.html', context)

@login_required
def mapa(request):
    """Vista del mapa interactivo"""
    pilares = Pilares.objects.all()
    
    datos = []
    for p in pilares:
        datos.append({
            'nombre': p.nombre,
            'alcaldia': p.alcaldia,
            'zona': p.zona,
            'modelo_equipo': p.modelo_equipo,  # ← NUEVO
            'lat': float(p.latitud),
            'lng': float(p.longitud),
            'pasta_termica': p.pasta_termica,
            'mouse_nuevos': p.mouse_nuevos,
            'teclados_nuevos': p.teclados_nuevos,
            'equipos_8gb': p.equipos_8gb,
            'total_equipos': p.total_equipos,
            'porcentaje_8gb': p.porcentaje_8gb,
        })
    
    datos_json = json.dumps(datos)
    
    return render(request, 'pilares/mapa.html', {'datos_pilares': datos_json})

@user_passes_test(es_admin)
@csrf_exempt
def editar_pilares(request, clave_id):
    """API para editar un PILARES"""
    if request.method == 'POST':
        try:
            pilares = Pilares.objects.get(clave_id=clave_id)
            data = json.loads(request.body)
            
            pilares.pasta_termica = data.get('pasta_termica', pilares.pasta_termica)
            pilares.mouse_nuevos = data.get('mouse_nuevos', pilares.mouse_nuevos)
            pilares.teclados_nuevos = data.get('teclados_nuevos', pilares.teclados_nuevos)
            pilares.equipos_8gb = data.get('equipos_8gb', pilares.equipos_8gb)
            pilares.total_equipos = data.get('total_equipos', pilares.total_equipos)
            pilares.equipos_inactivos = data.get('equipos_inactivos', pilares.equipos_inactivos)
            pilares.modelo_equipo = data.get('modelo_equipo', pilares.modelo_equipo)
            pilares.observaciones = data.get('observaciones', pilares.observaciones)
            pilares.save()
            
            return JsonResponse({'success': True, 'message': 'Actualizado correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# ============ SEGUIMIENTO DE EQUIPOS NO FUNCIONALES ============
def seguimiento(request):
    """Vista para el seguimiento de equipos no funcionales"""
    equipos = EquipoNoFuncional.objects.select_related('pilares').all()
    
    # Estadísticas para el dashboard de seguimiento
    total_equipos = equipos.count()
    pendientes = equipos.filter(estado='PENDIENTE').count()
    reparados = equipos.filter(estado='REPARADO').count()
    desechados = equipos.filter(estado='DESECHADO').count()
    
    # Datos para la tabla (con nombres de PILARES)
    equipos_data = []
    for e in equipos:
        equipos_data.append({
            'id': e.id,
            'pilares_id': e.pilares.clave_id,
            'pilares_nombre': e.pilares.nombre,
            'equipo': e.equipo,
            'tipo_falla': e.get_tipo_falla_display(),
            'tipo_falla_valor': e.tipo_falla,
            'descripcion': e.descripcion,
            'mouse': e.mouse,
            'teclado': e.teclado,
            'monitor': e.monitor,
            'estado': e.get_estado_display(),
            'estado_valor': e.estado,
            'fecha_reporte': e.fecha_reporte.strftime('%Y-%m-%d'),
            'fecha_solucion': e.fecha_solucion.strftime('%Y-%m-%d') if e.fecha_solucion else '',
            'observaciones': e.observaciones,
        })
    
    # Convertir a JSON de forma segura
    try:
        equipos_json = json.dumps(equipos_data, ensure_ascii=False)
    except Exception as e:
        equipos_json = json.dumps([])
        print(f"Error al serializar: {e}")
    
    # Lista de PILARES para el selector
    pilares_list = Pilares.objects.all().order_by('nombre')
    
    context = {
        'equipos_data_json': equipos_json,
        'total_equipos': total_equipos,
        'pendientes': pendientes,
        'reparados': reparados,
        'desechados': desechados,
        'pilares_list': pilares_list,
    }
    
    return render(request, 'pilares/seguimiento.html', context)

# ============ API AGREGAR EQUIPO ============
@csrf_exempt
def agregar_equipo_no_funcional(request):
    """API para agregar un equipo no funcional"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pilares = Pilares.objects.get(clave_id=data.get('pilares_id'))
            
            equipo = EquipoNoFuncional.objects.create(
                pilares=pilares,
                equipo=data.get('equipo', ''),
                tipo_falla=data.get('tipo_falla', 'OTRO'),
                descripcion=data.get('descripcion', ''),
                mouse=data.get('mouse', False),
                teclado=data.get('teclado', False),
                monitor=data.get('monitor', False),
                observaciones=data.get('observaciones', ''),
            )
            
            return JsonResponse({'success': True, 'id': equipo.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# ============ API EDITAR EQUIPO ============
@csrf_exempt
def editar_equipo_no_funcional(request, equipo_id):
    """API para editar un equipo no funcional"""
    if request.method == 'POST':
        try:
            equipo = EquipoNoFuncional.objects.get(id=equipo_id)
            data = json.loads(request.body)
            
            equipo.equipo = data.get('equipo', equipo.equipo)
            equipo.tipo_falla = data.get('tipo_falla', equipo.tipo_falla)
            equipo.descripcion = data.get('descripcion', equipo.descripcion)
            equipo.mouse = data.get('mouse', equipo.mouse)
            equipo.teclado = data.get('teclado', equipo.teclado)
            equipo.monitor = data.get('monitor', equipo.monitor)
            equipo.estado = data.get('estado', equipo.estado)
            equipo.observaciones = data.get('observaciones', equipo.observaciones)
            
            if data.get('fecha_solucion'):
                from datetime import datetime
                equipo.fecha_solucion = datetime.strptime(data['fecha_solucion'], '%Y-%m-%d').date()
            
            equipo.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@user_passes_test(es_admin)
def generar_reporte_pdf(request):
    """Genera un reporte PDF con todas las estadísticas de PILARES"""
    
    # Crear objeto de respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_pilares_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Configurar documento con márgenes más amplios
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           rightMargin=1.5*cm, leftMargin=1.5*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=20, textColor=colors.HexColor('#2c3e50'))
    heading_style = ParagraphStyle('HeadingStyle', parent=styles['Heading2'], fontSize=12, spaceAfter=10, textColor=colors.HexColor('#3498db'))
    normal_style = styles['Normal']
    
    # Estilo para texto con ajuste de línea
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10)
    
    # Elementos del PDF
    elementos = []
    
    # Título
    elementos.append(Paragraph("📊 Reporte de Mantenimiento - PILARES CDMX", title_style))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", normal_style))
    elementos.append(Spacer(1, 20))
    
    # ============ ESTADÍSTICAS GENERALES ============
    elementos.append(Paragraph("📈 Resumen General", heading_style))
    
    pilares = Pilares.objects.all()
    total_pilares = pilares.count()
    total_equipos_8gb = sum(p.equipos_8gb for p in pilares)
    total_equipos = sum(p.total_equipos for p in pilares)
    porcentaje_general = round((total_equipos_8gb / total_equipos * 100), 1) if total_equipos > 0 else 0
    total_mouse = sum(p.mouse_nuevos for p in pilares)
    total_teclados = sum(p.teclados_nuevos for p in pilares)
    total_pasta = round(sum(p.pasta_termica for p in pilares) / total_pilares, 1) if total_pilares > 0 else 0
    
    datos_generales = [
        [Paragraph('<b>Indicador</b>', cell_style), Paragraph('<b>Valor</b>', cell_style)],
        ['Total PILARES', str(total_pilares)],
        ['Equipos con 8GB RAM', str(total_equipos_8gb)],
        ['Total de equipos', str(total_equipos)],
        ['% Equipos 8GB', f"{porcentaje_general}%"],
        ['Mouse nuevos entregados', str(total_mouse)],
        ['Teclados nuevos entregados', str(total_teclados)],
        ['Promedio % Pasta térmica', f"{total_pasta}%"],
    ]
    
    tabla_general = Table(datos_generales, colWidths=[5*cm, 5*cm])
    tabla_general.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elementos.append(tabla_general)
    elementos.append(Spacer(1, 20))
    
    # ============ ESTADÍSTICAS POR ZONA ============
    elementos.append(Paragraph("🗺️ Estadísticas por Zona", heading_style))
    
    zonas = ['NORTE', 'SUR', 'PONIENTE', 'CENTRO', 'ORIENTE', 'SUR PONIENTE']
    datos_zonas = [[
        Paragraph('<b>Zona</b>', cell_style), 
        Paragraph('<b>PILARES</b>', cell_style),
        Paragraph('<b>Equipos 8GB</b>', cell_style), 
        Paragraph('<b>Total Equipos</b>', cell_style),
        Paragraph('<b>% 8GB</b>', cell_style), 
        Paragraph('<b>% Pasta</b>', cell_style)
    ]]
    
    for zona in zonas:
        pilares_zona = [p for p in pilares if p.zona == zona]
        count = len(pilares_zona)
        equipos_8gb = sum(p.equipos_8gb for p in pilares_zona)
        total_eq = sum(p.total_equipos for p in pilares_zona)
        porcentaje = round((equipos_8gb / total_eq * 100), 1) if total_eq > 0 else 0
        pasta_prom = round(sum(p.pasta_termica for p in pilares_zona) / count, 1) if count > 0 else 0
        datos_zonas.append([
            zona, 
            str(count), 
            str(equipos_8gb), 
            str(total_eq), 
            f"{porcentaje}%", 
            f"{pasta_prom}%"
        ])
    
    tabla_zonas = Table(datos_zonas, colWidths=[2.5*cm, 2*cm, 2.5*cm, 2.5*cm, 2*cm, 2.5*cm])
    tabla_zonas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elementos.append(tabla_zonas)
    elementos.append(Spacer(1, 20))
    
    # ============ TOP 10 PILARES ============
    elementos.append(Paragraph("🏆 Top 10 PILARES con mejor % de equipos 8GB", heading_style))
    
    pilares_ordenados = sorted(pilares, key=lambda p: p.porcentaje_8gb, reverse=True)[:10]
    datos_top = [[
        Paragraph('<b>CLAVE_ID</b>', cell_style),
        Paragraph('<b>PILARES</b>', cell_style),
        Paragraph('<b>Alcaldía</b>', cell_style),
        Paragraph('<b>8GB/TOTAL</b>', cell_style),
        Paragraph('<b>% 8GB</b>', cell_style),
        Paragraph('<b>% Pasta</b>', cell_style)
    ]]
    
    for p in pilares_ordenados:
        # Truncar textos largos y usar Paragraph para ajuste automático
        nombre = p.nombre[:25] + '...' if len(p.nombre) > 25 else p.nombre
        alcaldia = p.alcaldia[:20] + '...' if len(p.alcaldia) > 20 else p.alcaldia
        datos_top.append([
            p.clave_id,
            Paragraph(nombre, cell_style),
            Paragraph(alcaldia, cell_style),
            f"{p.equipos_8gb}/{p.total_equipos}",
            f"{p.porcentaje_8gb}%",
            f"{p.pasta_termica}%"
        ])
    
    tabla_top = Table(datos_top, colWidths=[2.5*cm, 4.5*cm, 3*cm, 2.5*cm, 1.8*cm, 1.8*cm])
    tabla_top.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e67e22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elementos.append(tabla_top)
    elementos.append(Spacer(1, 20))
    
    # ============ PILARES CRÍTICOS ============
    elementos.append(Paragraph("⚠️ PILARES que requieren atención (% 8GB menor a 50%)", heading_style))
    
    pilares_criticos = [p for p in pilares if p.porcentaje_8gb < 50][:15]
    if pilares_criticos:
        datos_criticos = [[
            Paragraph('<b>CLAVE_ID</b>', cell_style),
            Paragraph('<b>PILARES</b>', cell_style),
            Paragraph('<b>Alcaldía</b>', cell_style),
            Paragraph('<b>8GB/TOTAL</b>', cell_style),
            Paragraph('<b>% 8GB</b>', cell_style)
        ]]
        for p in pilares_criticos:
            nombre = p.nombre[:25] + '...' if len(p.nombre) > 25 else p.nombre
            alcaldia = p.alcaldia[:20] + '...' if len(p.alcaldia) > 20 else p.alcaldia
            datos_criticos.append([
                p.clave_id,
                Paragraph(nombre, cell_style),
                Paragraph(alcaldia, cell_style),
                f"{p.equipos_8gb}/{p.total_equipos}",
                f"{p.porcentaje_8gb}%"
            ])
        tabla_criticos = Table(datos_criticos, colWidths=[2.5*cm, 4.5*cm, 3*cm, 2.5*cm, 2*cm])
        tabla_criticos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elementos.append(tabla_criticos)
    else:
        elementos.append(Paragraph("✅ ¡Excelente! No hay PILARES con porcentaje crítico.", normal_style))
    
    # Generar PDF
    doc.build(elementos)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

# ============ INCIDENCIAS ============
def incidencias(request):
    """Vista de la pestaña de incidencias"""
    from .models import ObservacionHistorial, Pilares
    pilares_list = Pilares.objects.all().order_by('nombre')
    return render(request, 'pilares/incidencias.html', {'pilares_list': pilares_list})

def api_incidencias_pilares(request, clave_id):
    """API para obtener las incidencias de un PILARES en JSON"""
    from .models import ObservacionHistorial
    historial = ObservacionHistorial.objects.filter(pilares__clave_id=clave_id).order_by('fecha_creacion')
    
    data = []
    for h in historial:
        data.append({
            'id': h.id,
            'fecha_creacion': h.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'tipo_cambio': h.tipo_cambio,
            'tipo_cambio_display': h.get_tipo_cambio_display(),
            'estado_nuevo': h.estado_nuevo,
            'estado_nuevo_display': h.get_estado_nuevo_display(),
            'observacion_anterior': h.observacion_anterior,
            'observacion_nueva': h.observacion_nueva,
        })
    
    return JsonResponse(data, safe=False)

def historial_general(request):
    """Vista del historial general de observaciones"""
    try:
        from .models import ObservacionHistorial
        historial = ObservacionHistorial.objects.select_related('pilares', 'usuario').all()[:100]
        total = ObservacionHistorial.objects.count()
    except:
        # Si el modelo no existe, mostrar vacío
        historial = []
        total = 0
    
    context = {
        'historial': historial,
        'total': total,
    }
    return render(request, 'pilares/historial_general.html', context)

# ============ CREAR INCIDENCIA ============
@csrf_exempt
@login_required
def crear_incidencia(request):
    """API para crear una nueva incidencia"""
    if request.method == 'POST':
        try:
            from .models import ObservacionHistorial, Pilares
            data = json.loads(request.body)
            
            pilares = Pilares.objects.get(clave_id=data.get('clave_id'))
            
            # Crear la incidencia
            incidencia = ObservacionHistorial.objects.create(
                pilares=pilares,
                usuario=request.user,
                observacion_nueva=data.get('observacion', ''),
                tipo_cambio='CREACION',
                estado_nuevo='ABIERTA',
                equipo_afectado=data.get('equipo', ''),
            )
            
            # Actualizar el estado del PILARES
            pilares.estado_incidencia = 'ABIERTA'
            pilares.observaciones = data.get('observacion', '')
            pilares.save()
            
            return JsonResponse({'success': True, 'id': incidencia.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# ============ ACTUALIZAR INCIDENCIA ============
@csrf_exempt
@login_required
def actualizar_incidencia(request, incidencia_id):
    """API para actualizar una incidencia existente"""
    if request.method == 'POST':
        try:
            from .models import ObservacionHistorial
            data = json.loads(request.body)
            
            incidencia = ObservacionHistorial.objects.get(id=incidencia_id)
            
            # Guardar la observación anterior
            incidencia.observacion_anterior = incidencia.observacion_nueva
            
            # Actualizar con los nuevos datos
            incidencia.observacion_nueva = data.get('observacion', incidencia.observacion_nueva)
            incidencia.estado_nuevo = data.get('estado', incidencia.estado_nuevo)
            incidencia.tipo_cambio = 'ACTUALIZACION'
            incidencia.save()
            
            # Actualizar el PILARES
            pilares = incidencia.pilares
            pilares.observaciones = incidencia.observacion_nueva
            pilares.estado_incidencia = incidencia.estado_nuevo
            pilares.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def reporte_incidencias_general_pdf(request):
    """Genera un PDF con TODAS las incidencias de todos los PILARES (formato mejorado)"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from .models import ObservacionHistorial
    import io
    from datetime import datetime
    
    historial = ObservacionHistorial.objects.select_related('pilares', 'usuario').order_by('-fecha_creacion')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_incidencias_general_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=1.5*cm, leftMargin=1.5*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=10, textColor=colors.HexColor('#2c3e50'))
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=15)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_LEFT)
    obs_style = ParagraphStyle('ObsStyle', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_LEFT)
    
    elementos = []
    
    elementos.append(Paragraph("📋 Reporte General de Incidencias", title_style))
    elementos.append(Paragraph("Sistema de Mantenimiento - PILARES CDMX", subtitle_style))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elementos.append(Spacer(1, 10))
    
    # Estadísticas
    total = historial.count()
    abiertas = historial.filter(estado_nuevo='ABIERTA').count()
    en_proceso = historial.filter(estado_nuevo='EN_PROCESO').count()
    solucionadas = historial.filter(estado_nuevo='SOLUCIONADA').count()
    cerradas = historial.filter(estado_nuevo='CERRADA').count()
    
    resumen_data = [
        ['Total', '🟡 Abiertas', '🔵 En proceso', '🟢 Solucionadas', '⚫ Cerradas'],
        [str(total), str(abiertas), str(en_proceso), str(solucionadas), str(cerradas)]
    ]
    
    tabla_resumen = Table(resumen_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 20))
    
    # ============ SI HAY DATOS ============
    if historial.exists():
        historial_list = list(historial)
        total_registros = len(historial_list)
        batch_size = 15
        
        for start_idx in range(0, total_registros, batch_size):
            batch = historial_list[start_idx:start_idx + batch_size]
            
            if start_idx > 0:
                elementos.append(PageBreak())
                elementos.append(Paragraph("📋 Reporte General de Incidencias (continuación)", title_style))
                elementos.append(Spacer(1, 10))
            
            # ============ CONSTRUIR TABLA ============
            table_data = [['#', 'Fecha', 'PILARES', 'Tipo', 'Estado', 'Observación']]
            
            for idx, h in enumerate(batch, start_idx + 1):
                estado = h.estado_nuevo or 'ABIERTA'
                color_fila = '#27ae60' if estado == 'SOLUCIONADA' else '#e74c3c' if estado == 'EN_PROCESO' else '#f39c12'
                obs = h.observacion_nueva or '-'
                obs_paragraph = Paragraph(obs, obs_style)
                
                table_data.append([
                    str(idx),
                    h.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                    Paragraph(h.pilares.nombre[:25], cell_style),
                    Paragraph(h.get_tipo_cambio_display(), cell_style),
                    Paragraph(f'<font color="{color_fila}">{h.get_estado_nuevo_display() or "Sin estado"}</font>', cell_style),
                    obs_paragraph,
                ])
            
            # ============ CREAR TABLA ============
            tabla = Table(table_data, colWidths=[0.6*cm, 3.5*cm, 4*cm, 2.5*cm, 2.5*cm, 5.5*cm])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#f9f9f9')]),
                ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ]))
            elementos.append(tabla)
            elementos.append(Spacer(1, 10))
        
        elementos.append(Spacer(1, 10))
        elementos.append(Paragraph("Reporte generado automáticamente por el sistema PILARES CDMX", styles['Normal']))
    
    # ============ SI NO HAY DATOS ============
    else:
        elementos.append(Paragraph("No hay incidencias registradas", styles['Normal']))
    
    doc.build(elementos)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response
# ============ REPORTE EXCEL GENERAL DE INCIDENCIAS ============
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def reporte_incidencias_general_excel(request):
    """Genera un Excel con TODAS las incidencias de todos los PILARES"""
    from .models import ObservacionHistorial
    from datetime import datetime
    
    historial = ObservacionHistorial.objects.select_related('pilares', 'usuario').order_by('-fecha_creacion')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_incidencias_general_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias"
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Encabezados
    headers = ['#', 'Fecha', 'PILARES', 'Alcaldía', 'Zona', 'Tipo', 'Estado', 'Observación']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for idx, h in enumerate(historial, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=h.fecha_creacion.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3, value=h.pilares.nombre)
        ws.cell(row=row, column=4, value=h.pilares.alcaldia)
        ws.cell(row=row, column=5, value=h.pilares.zona)
        ws.cell(row=row, column=6, value=h.get_tipo_cambio_display())
        ws.cell(row=row, column=7, value=h.get_estado_nuevo_display() or 'Sin estado')
        ws.cell(row=row, column=8, value=h.observacion_nueva or '-')
        
        # Aplicar alignment a todas las celdas de la fila
        for col in range(1, 9):
            ws.cell(row=row, column=col).alignment = cell_alignment
    
    # Ajustar ancho de columnas
    column_widths = [4, 18, 30, 25, 15, 15, 20, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    wb.save(response)
    return response

